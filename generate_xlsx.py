import camelot
from openpyxl import Workbook
from operator import attrgetter

PDF   = (
    "https://www.jiit.ac.in/sites/default/files/B.TECH_.%20I%20Sem%20Odd%202025%20Wish%20Town.pdf"
)

PDF = "./B.TECH. I Yr(II SEMESTER) TIMETABLE EVEN SEMESTER 2026, JIIT-128.pdf"
PAGES = "all"

tables = camelot.read_pdf(
    PDF,
    pages=PAGES,
    flavor="lattice",
    line_scale=40,
    copy_text=None,
)

tables = sorted(tables, key=attrgetter("page"))

wb = Workbook()
ws = wb.active
ws.title = "Timetable"

excel_row = 1  

for table in tables:
    table.set_border()
    nrows, ncols = len(table.rows), len(table.cols)
    visited = set()

    for r in range(nrows):
        c = 0
        while c < ncols:
            if (r, c) in visited:
                c += 1
                continue

            cell = table.cells[r][c]

            # Determine span
            c_end = c
            while c_end < ncols - 1 and not table.cells[r][c_end].right:
                c_end += 1

            r_end = r
            while r_end < nrows - 1 and not table.cells[r_end][c].bottom:
                r_end += 1

            parts = []
            for rr in range(r, r_end + 1):
                for cc in range(c, c_end + 1):
                    t = table.cells[rr][cc].text.strip()
                    if t:
                        parts.append(t)
                    visited.add((rr, cc))

            # Unique lines only
            seen = set()
            ordered_unique = []
            for p in parts:
                if p not in seen:
                    seen.add(p)
                    ordered_unique.append(p)

            txt = "\n".join(ordered_unique).strip()

            if txt:
                ws.cell(row=excel_row + r, column=c + 1, value=txt)

            # Merge cells if needed
            if r_end > r or c_end > c:
                ws.merge_cells(
                    start_row=excel_row + r,
                    start_column=c + 1,
                    end_row=excel_row + r_end,
                    end_column=c_end + 1,
                )
            c = c_end + 1
    excel_row += nrows + 2

wb.save("timetable.xlsx")

